# _*_ coding:utf-8 _*_
"""
@Software  ：api_learning
@FileName  ：handle_excel.py
@Date      ：2023/11/23 17:29
@Author    ：ChenGH
"""
import openpyxl


class HandleExcel(object):
    titles = ('id', 'title', 'data', 'expected')

    def __init__(self, filename, sheet_name=None):
        self.filename = filename
        if sheet_name is None:
            self.sheet_name = 'Sheet1'
        else:
            self.sheet_name = sheet_name

    def read_data(self) -> list:
        wb = openpyxl.load_workbook(self.filename)
        ws = wb[self.sheet_name]
        cases = []
        titles = []

        for i, row in enumerate(ws.rows):
            if i == 0:
                titles = [cell.value for cell in row]
                # for cell in row:
                #     titles.append(cell.value)
            else:
                cases.append(dict(zip(titles, [cell.value for cell in row])))
        return cases

    def write_data(self, row, column, value) -> None:
        wb = openpyxl.load_workbook(self.filename)
        ws = wb[self.sheet_name]
        ws.cell(row, column, value)
        wb.save(self.filename)
        return None


if __name__ == '__main__':
    file = r'E:\安装包-勿删&总结\api-auto-test\api_learning\data\apicases.xlsx'
    handle = HandleExcel(file, sheet_name='login')
    for case in handle.read_data():
        print(case)
    # handle.write_data(2, 8, 'pass')  # 向第二行、第八列写入数据
